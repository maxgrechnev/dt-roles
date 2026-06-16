import string
import requests
import urllib3
import argparse
import warnings
import logging
import sys
import copy
from openpyxl import load_workbook, Workbook
from enum import Enum

MAX_NAME_LENGTH = 100
NAME_FIELDS_DELIMITER = ' - '
GROUP_NAME_PATTERN = '{role} - {mz}'
SERVER_URL = 'https://{hostname}/'
ENV_URL_PATH = 'e/{env_id}/'
CLUSTER_API_PATH_USERS = 'api/v1.0/onpremise/users'
CLUSTER_API_PATH_GROUPS = 'api/v1.0/onpremise/groups'
CLUSTER_API_PATH_GROUPS_MZ = 'api/v1.0/onpremise/groups/managementZones'
ENV_API_PATH_MZ = 'api/config/v1/managementZones'
HEADERS = {
    'Authorization': 'Api-Token {token}',
    'Content-Type': 'application/json; charset=utf8',
}
ROLE_JSON_TEMPLATE = {
    "cluster": False,
    "env": {
        "VIEWER": False,
        "MANAGE_SETTINGS": False,
        "AGENT_INSTALL": False,
        "LOG_VIEWER": False,
        "VIEW_SENSITIVE_REQUEST_DATA": False,
        "CONFIGURE_REQUEST_CAPTURE_DATA": False,
        "REPLAY_SESSION_DATA": False
    },
    "mz": {
        "MANAGE_SETTINGS": False,
        "VIEWER": False,
        "VIEW_SENSITIVE_REQUEST_DATA": False,
        "LOG_VIEWER": False
    }
}
GROUP_JSON_TEMPLATE = {
    "isClusterAdminGroup": False,
    "name": "string",
    "accessRight": {}
}
GROUP_MZ_JSON_TEMPLATE = {
    "groupId": "string",
    "mzPermissionsPerEnvironment": [
        {
            "environmentUuid": "string",
            "mzPermissions": []
        }
    ]
}
MZ_JSON_TEMPLATE = {
    "name": "string",
    "rules": [
        {
            "type": "HOST",
            "enabled": True,
            "propagationTypes": [
                "HOST_TO_PROCESS_GROUP_INSTANCE"
            ],
            "conditions": [
                {
                    "key": {
                        "attribute": "HOST_TAGS"
                    },
                    "comparisonInfo": {
                        "type": "TAG",
                        "operator": "EQUALS",
                        "value": {
                            "context": "CONTEXTLESS",
                            "key": "string"
                        },
                        "negate": False
                    }
                }
            ]
        },
        {
            "type": "PROCESS_GROUP",
            "enabled": True,
            "propagationTypes": [
                "PROCESS_GROUP_TO_HOST",
                "PROCESS_GROUP_TO_SERVICE"
            ],
            "conditions": [
                {
                    "key": {
                        "attribute": "PROCESS_GROUP_TAGS"
                    },
                    "comparisonInfo": {
                        "type": "TAG",
                        "operator": "EQUALS",
                        "value": {
                            "context": "CONTEXTLESS",
                            "key": "string",
                            "value": "string"
                        },
                        "negate": False
                    }
                }
            ]
        },
        {
            "type": "SERVICE",
            "enabled": True,
            "propagationTypes": [
                "SERVICE_TO_HOST_LIKE",
                "SERVICE_TO_PROCESS_GROUP_LIKE"
            ],
            "conditions": [
                {
                    "key": {
                        "attribute": "SERVICE_TAGS"
                    },
                    "comparisonInfo": {
                        "type": "TAG",
                        "operator": "EQUALS",
                        "value": {
                            "context": "CONTEXTLESS",
                            "key": "string",
                            "value": "string"
                        },
                        "negate": False
                    }
                }
            ]
        }
    ]
}


def parse_args():
    arg_parser = argparse.ArgumentParser(description='Automate Dynatrace user roles creation and assignments')
    arg_parser.add_argument('--cluster-hostname', dest='cluster_hostname', required=True,
                            help='Dynatrace cluster hostname')
    arg_parser.add_argument('--cluster-token', dest='cluster_token', required=True,
                            help='Dynatrace API cluster token')
    arg_parser.add_argument('--env-id', dest='env_id', required=True, help='Dynatrace environment ID')
    arg_parser.add_argument('--env-token', dest='env_token', required=True, help='Dynatrace API environment token')
    arg_parser.add_argument('--file', dest='file', required=True, help='Input Excel file')
    arg_parser.add_argument('--users-sheet', dest='users_sheet', required=True,
                            help='Users list sheet name in the specified Excel file')
    arg_parser.add_argument('--roles-sheet', dest='roles_sheet', required=True,
                            help='Role definitions sheet name in the specified Excel file')
    arg_parser.add_argument('--apply-changes', dest='apply_changes', required=False, action='store_true',
                            help='Apply changes to Dynatrace configuration_import_export (by default run in test mode)')
    return arg_parser.parse_args()


class DtRolesError(Exception):
    """Wrapper for Dynatrace roles exceptions"""
    pass


class DtAPIType(Enum):
    """Enumeration of Dynatrace API types"""
    CLUSTER = 'cluster'
    ENV = 'env'


class DtRoleLevel(Enum):
    """Enumeration of Dynatrace role levels of permissions"""
    CLUSTER = 'cluster'
    ENV = 'env'
    MZ = 'mz'


class DtObjectType(Enum):
    """Enumeration of Dynatrace configuration_import_export objects managed by the tool"""
    GROUP = 'group'
    MZ = 'mz'


class HTTPRequestType(Enum):
    """Enumeration of HTTP request methods used by the tool"""
    GET = 'GET'
    POST = 'POST'
    PUT = 'PUT'


class DtMzTagComparisonOperator(Enum):
    """Enumeration of Dynatrace API comparison operators for tag filter in management zone"""
    KEY_VALUE = 'EQUALS'
    ONLY_KEY = 'TAG_KEY_EQUALS'


class DynatraceRoles:
    """Class for managing Dynatrace user roles"""

    def __init__(self, cluster_hostname, cluster_token, env_id, env_token):
        self.env_id = env_id
        self.input_file = ''
        self.users_sheet = ''
        self.roles_sheet = ''
        self.workbook = Workbook()
        self.roles = {}
        self.groups_created = []
        self.mz_created = []
        self.permissions_updated = []
        self.users_added = []
        self.url = {
            DtAPIType.CLUSTER: SERVER_URL.format(hostname=cluster_hostname),
            DtAPIType.ENV: SERVER_URL.format(hostname=cluster_hostname) + ENV_URL_PATH.format(env_id=env_id),
        }
        self.headers = {
            DtAPIType.CLUSTER: HEADERS.copy(),
            DtAPIType.ENV: HEADERS.copy(),
        }
        self.headers[DtAPIType.CLUSTER]['Authorization'] = HEADERS['Authorization'].format(token=cluster_token)
        self.headers[DtAPIType.ENV]['Authorization'] = HEADERS['Authorization'].format(token=env_token)
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        self.log = logging.getLogger('dt-roles')
        self._configure_logging()

    def _configure_logging(self):
        """Configure default logging settings"""
        self.log.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s')
        stdout_handler = logging.StreamHandler(sys.stdout)
        stdout_handler.setFormatter(formatter)
        self.log.addHandler(stdout_handler)

    def _load_role_definitions(self, roles_sheet):
        """Load Dynatrace role definitions from the specified Excel sheet"""
        self.roles_sheet = roles_sheet
        if roles_sheet not in self.workbook:
            raise DtRolesError(
                'Roles sheet "{}" does not exist in the file "{}"'.format(roles_sheet, self.input_file))

        sheet_values = self.workbook[roles_sheet].values
        role_level_header = sheet_values.__next__()[1:]
        role_perm_header = sheet_values.__next__()[1:]
        header_length = len(role_perm_header)

        # Loop through roles on the sheet
        for row in sheet_values:
            role_name = row[0]
            if not role_name:
                # Empty first cell: end of roles list
                break
            role = copy.deepcopy(ROLE_JSON_TEMPLATE)

            # Loop through role permissions in the row
            for i, value in enumerate(row[1:header_length]):
                # If not empty cell: permission is assigned
                if value:
                    role_level = role_level_header[i]
                    role_perm = role_perm_header[i]

                    if role_level == DtRoleLevel.CLUSTER.value:
                        role[role_level] = True

                    elif role_level in [DtRoleLevel.ENV.value, DtRoleLevel.MZ.value]:
                        if role_perm in role[role_level]:
                            role[role_level][role_perm] = True
                        else:
                            raise DtRolesError(
                                'Unsupported permission "{}" specified '.format(role_perm) +
                                'on the sheet "{}" in the file "{}": it should be one of the following: {}'.format(
                                        roles_sheet, self.input_file, str(list(role[role_level].keys()))))
                    else:
                        raise DtRolesError(
                            'Unsupported role level "{}" specified '.format(role_level) +
                            'on the sheet "{}" in the file "{}": it should be one of the following: {}'.format(
                                roles_sheet, self.input_file, str([i.value for i in DtRoleLevel])))

            self.roles[role_name] = role

        self._check_role_definitions()

    def _check_role_definitions(self):
        """Check loaded Dynatrace role definitions for correctness"""
        for role_key in self.roles.keys():
            role = self.roles[role_key]
            for perm_type in ['env', 'mz']:
                env_viewer = role[perm_type]['VIEWER']
                rest_permissions = False
                for perm_key in role[perm_type].keys():
                    if perm_key != 'VIEWER' and role[perm_type][perm_key]:
                        rest_permissions = True
                if not env_viewer and rest_permissions:
                    raise DtRolesError(
                        'Role definition "{}" in the file "{}" '.format(role_key, self.input_file) +
                        'sheet "{}" is incorrect: access environment permission (VIEWER) '.format(self.roles_sheet) +
                        'must be enabled when at least any other permission is enabled')

    def _get_role_level(self, role):
        """Determine Dynatrace role level from it's permissions specified in the role definition"""
        if role not in self.roles:
            raise DtRolesError(
                'Role "{}" is not defined on the roles sheet "{}" in the file "{}": define it and try again'.format(
                    role, self.roles_sheet, self.input_file))

        role_def = self.roles[role]
        if role_def['cluster']:
            return DtRoleLevel.CLUSTER
        else:
            mz_level = False
            # Level is MZ if at least one MZ permission
            for mz_perm in role_def['mz'].values():
                if mz_perm:
                    mz_level = True
            if mz_level:
                return DtRoleLevel.MZ
            else:
                return DtRoleLevel.ENV

    def _get_api_response(self, api_type, api_path, request_type, params=None, body=None):
        """Perform specified Dynatrace API request and get the response"""
        connect_url = self.url[api_type] + api_path
        headers = self.headers[api_type]
        response = None
        try:
            if request_type == HTTPRequestType.GET:
                response = requests.get(url=connect_url, params=params, headers=headers, json=body, verify=False)
            elif request_type == HTTPRequestType.POST:
                response = requests.post(url=connect_url, params=params, headers=headers, json=body, verify=False)
            elif request_type == HTTPRequestType.PUT:
                response = requests.put(url=connect_url, params=params, headers=headers, json=body, verify=False)

            response.raise_for_status()
            return response

        except requests.exceptions.ConnectionError as e:
            raise DtRolesError(str(e))

        except requests.exceptions.HTTPError as e:
            raise DtRolesError('{}: {}'.format(e, response.text))

    def _get_current_configuration(self):
        """Get existing Dynatrace configuration_import_export objects: users, groups and management zones"""
        response = self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_GROUPS, HTTPRequestType.GET)
        self.groups = {
            group['name']: group
            for group in response.json()
        }
        response = self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_GROUPS_MZ, HTTPRequestType.GET)
        self.groups_mz = {
            group['groupId']: group
            for group in response.json()
        }
        response = self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_USERS, HTTPRequestType.GET)
        self.users = {
            user['id']: user
            for user in response.json()
        }
        response = self._get_api_response(DtAPIType.ENV, ENV_API_PATH_MZ, HTTPRequestType.GET)
        self.zones = {
            zone['name']: zone
            for zone in response.json()['values']
        }

    def _open_users_sheet(self, users_sheet):
        """Open users sheet from Excel file, read and check table's header, return sheet rows iterator"""
        self.users_sheet = users_sheet
        if users_sheet not in self.workbook:
            raise DtRolesError(
                'Users sheet "{}" does not exist in the file "{}"'.format(users_sheet, self.input_file))

        sheet_iterator = enumerate(self.workbook[users_sheet].values)
        _, header = sheet_iterator.__next__()
        if len(header) < 5:
            raise DtRolesError(
                'Users sheet "{}" in the file "{}" should contain at least 5 columns'.format(users_sheet,
                                                                                             self.input_file))
        return sheet_iterator

    @staticmethod
    def _get_longest_string_index(str_list):
        """Get the longest string index from a list of strings"""
        longest = 0
        for i in range(len(str_list)):
            if len(str_list[i]) > len(str_list[longest]):
                longest = i
        return longest

    def _get_name_from_row(self, name_type, role_name, role_level, row_number, mz_name_parts):
        """Construct a name of the specified Dynatrace object using concatenation of name parts in a row"""
        if role_level == DtRoleLevel.MZ:
            mz_name_parts = list(mz_name_parts)
            # Try to check for name oversizing only twice
            for attempt in [1, 2]:
                name = NAME_FIELDS_DELIMITER.join(mz_name_parts)
                if name_type == DtObjectType.GROUP:
                    # Add role name to MZ name
                    name = GROUP_NAME_PATTERN.format(mz=name, role=role_name)

                # Check for oversizing
                oversize = len(name) - MAX_NAME_LENGTH
                if oversize > 0:
                    if attempt == 1:
                        # Cut the longest MZ name part and try again (continue to loop attempt 2)
                        longest = self._get_longest_string_index(mz_name_parts)
                        mz_name_parts[longest] = mz_name_parts[longest][:-oversize]
                    else:
                        raise DtRolesError('Cutting the longest MZ part in row {} has '.format(row_number) +
                                           'no effect: {} name still exceeds the limit of '.format(name_type.value) +
                                           '{} symbols: "{}". '.format(MAX_NAME_LENGTH, name) +
                                           'Try to use less MZ name parts or shorter role names.')
                else:
                    return name
        elif name_type == DtObjectType.GROUP:
            # Group name = role name for cluster and env level roles
            return role_name
        else:
            return None

    def _mz_already_exists(self, mz_name, tag_key, tag_value):
        """Check whether the specified management zone exists in Dynatrace environment at the moment"""
        if mz_name in self.zones:
            mz = self.zones[mz_name]
            if 'rules' in mz.keys():
                # This means that MZ was created during this execution and contains the whole JSON object.
                # So we can compare and check tags. For previously existing zones we collect only id and name.
                mz_tag_key = mz['rules'][0]['conditions'][0]['comparisonInfo']['value']
                if mz_tag_key['key'] != tag_key or mz_tag_key['value'] != tag_value:
                    raise DtRolesError(
                        'Some rows on the users sheet "{}" have conflicts: '.format(self.users_sheet) +
                        'different tags specified for the same MZ "{}"'.format(mz_name))
            return True
        else:
            return False

    @staticmethod
    def _get_mz_json_from_template(mz_name, comparison_operator, tag_key, tag_value):
        """Create a new management zone JSON object from the template and fill in filtering tag key-value pairs"""
        mz_json = copy.deepcopy(MZ_JSON_TEMPLATE)
        mz_json['name'] = mz_name
        for rule in mz_json['rules']:
            rule['conditions'][0]['comparisonInfo']['operator'] = comparison_operator.value
            rule['conditions'][0]['comparisonInfo']['value']['key'] = tag_key
            if comparison_operator == DtMzTagComparisonOperator.KEY_VALUE:
                rule['conditions'][0]['comparisonInfo']['value']['value'] = tag_value
        return mz_json

    def _create_mz_if_needed(self, mz_name, tag_key, tag_value, apply_changes=False):
        """Create a new management zone in Dynatrace environment if it does not exist"""
        if self._mz_already_exists(mz_name, tag_key, tag_value):
            self.log.info('MZ already exists: "{}"'.format(mz_name))
            return self.zones[mz_name]['id']
        else:
            if tag_value:
                comparison_operator = DtMzTagComparisonOperator.KEY_VALUE
            else:
                comparison_operator = DtMzTagComparisonOperator.ONLY_KEY

            mz_json = self._get_mz_json_from_template(mz_name, comparison_operator, tag_key, tag_value)
            if apply_changes:
                response = self._get_api_response(DtAPIType.ENV, ENV_API_PATH_MZ, HTTPRequestType.POST, body=mz_json)
                mz_json = response.json()
            else:
                # Fake MZ ID in test run
                mz_json['id'] = mz_name
            self.log.warning('MZ created: "{}"'.format(mz_name))
            self.zones[mz_name] = mz_json
            self.mz_created.append(mz_name)
            return mz_json['id']

    @staticmethod
    def _update_env_permissions_for_group(group_json, permissions, env_id):
        """Update the user group JSON object according to the specified environment level permissions.
        Return a tuple: (new JSON object, whether it changed - True/False).
        """
        perm_changed = False
        for perm_key in permissions.keys():
            if permissions[perm_key]:
                # Add permission if absent
                if perm_key not in group_json['accessRight'].keys():
                    group_json['accessRight'][perm_key] = []
                if env_id not in group_json['accessRight'][perm_key]:
                    group_json['accessRight'][perm_key].append(env_id)
                    perm_changed = True
            else:
                # Remove permission if assigned
                if perm_key in group_json['accessRight'].keys() and env_id in group_json['accessRight'][perm_key]:
                    group_json['accessRight'][perm_key].remove(env_id)
                    perm_changed = True
        return group_json, perm_changed

    @staticmethod
    def _update_mz_permissions_for_group(group_mz_json, permissions, env_id, mz_id):
        """Update the user group JSON object according to the specified management zone level permissions.
        Return a tuple: (new JSON object, whether it changed - True/False).
        """
        perm_changed = False
        for env_perms in group_mz_json['mzPermissionsPerEnvironment']:
            if env_perms['environmentUuid'] == env_id:
                for mz_perms in env_perms['mzPermissions']:
                    if mz_perms['mzId'] == mz_id:
                        for perm_key in permissions.keys():
                            if permissions[perm_key]:
                                if perm_key not in mz_perms['permissions']:
                                    mz_perms['permissions'].append(perm_key)
                                    perm_changed = True
                            else:
                                if perm_key in mz_perms['permissions']:
                                    mz_perms['permissions'].remove(perm_key)
                                    perm_changed = True
                        return group_mz_json, perm_changed
        return group_mz_json, perm_changed

    def _create_group(self, group_name, role, role_level, env_id=None, mz_id=None, apply_changes=False):
        """Create a new user group in Dynatrace cluster"""
        group_json = copy.deepcopy(GROUP_JSON_TEMPLATE)
        group_json['name'] = group_name

        if role_level == DtRoleLevel.CLUSTER:
            group_json['isClusterAdminGroup'] = True
        else:
            env_permissions = self.roles[role]['env']
            group_json, _ = self._update_env_permissions_for_group(group_json, env_permissions, env_id)

        if apply_changes:
            response = self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_GROUPS, HTTPRequestType.POST,
                                              body=group_json)
            group_json = response.json()
        else:
            # Fake group ID in test run
            group_json['id'] = group_name

        group_id = group_json['id']
        self.groups[group_name] = group_json

        group_mz_json = copy.deepcopy(GROUP_MZ_JSON_TEMPLATE)
        group_mz_json['groupId'] = group_id
        group_mz_json['mzPermissionsPerEnvironment'][0]['environmentUuid'] = env_id

        if role_level == DtRoleLevel.MZ:
            mz_permissions = self.roles[role]['mz']
            group_mz_json['mzPermissionsPerEnvironment'][0]['mzPermissions'].append({
                'mzId': mz_id,
                'permissions': []
            })
            group_mz_json, _ = self._update_mz_permissions_for_group(group_mz_json, mz_permissions, env_id, mz_id)
            if apply_changes:
                self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_GROUPS_MZ, HTTPRequestType.PUT, body=group_mz_json)

        self.log.warning('User group created: "{}"'.format(group_name))
        self.groups_mz[group_id] = group_mz_json
        self.groups_created.append(group_name)
        return group_id

    def _update_group_if_needed(self, group_name, role, role_level, env_id=None, mz_id=None, apply_changes=False):
        """Update the specified user group permissions in Dynatrace cluster if it differs from the role definition"""
        group_json = self.groups[group_name]
        group_id = group_json['id']
        group_mz_json = self.groups_mz[group_id]
        cluster_perm_changed = False
        env_perm_changed = False
        mz_perm_changed = False

        if role_level == DtRoleLevel.CLUSTER:
            if not group_json['isClusterAdminGroup']:
                group_json['isClusterAdminGroup'] = True
                cluster_perm_changed = True
        else:
            if group_json['isClusterAdminGroup']:
                group_json['isClusterAdminGroup'] = False
                cluster_perm_changed = True

            env_permissions = self.roles[role]['env']
            group_json, env_perm_changed = self._update_env_permissions_for_group(group_json, env_permissions, env_id)

            if role_level == DtRoleLevel.MZ:
                mz_permissions = self.roles[role]['mz']
                group_mz_json, mz_perm_changed = self._update_mz_permissions_for_group(group_mz_json, mz_permissions,
                                                                                       env_id, mz_id)

        if cluster_perm_changed or env_perm_changed or mz_perm_changed:
            if cluster_perm_changed or env_perm_changed:
                if apply_changes:
                    response = self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_GROUPS, HTTPRequestType.PUT,
                                                      body=group_json)
                    group_json = response.json()
                self.log.warning('User group cluster/environment permissions updated: "{}"'.format(group_name))
                self.groups[group_name] = group_json

            if mz_perm_changed:
                if apply_changes:
                    self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_GROUPS_MZ, HTTPRequestType.PUT,
                                           body=group_mz_json)
                self.log.warning('User group MZ permissions updated: "{}"'.format(group_name))
                self.groups_mz[group_id] = group_mz_json

            self.permissions_updated.append(group_name)
        else:
            self.log.info('User group already exists with the same permissions: "{}"'.format(group_name))
        return group_id

    def _add_user_to_group_if_needed(self, user_id, group_id, group_name, apply_changes=False):
        """Add the user to the specified group in Dynatrace cluster if it is not in the group"""
        if user_id not in self.users:
            raise DtRolesError('User "{}" does not exist in Dynatrace: create the user and try again'.format(user_id))
        user_json = self.users[user_id]

        if group_id in user_json['groups']:
            self.log.info('User "{}" is already in the group "{}"'.format(user_id, group_name))
        else:
            user_json['groups'].append(group_id)
            if apply_changes:
                response = self._get_api_response(DtAPIType.CLUSTER, CLUSTER_API_PATH_USERS, HTTPRequestType.PUT,
                                                  body=user_json)
                user_json = response.json()
            self.log.warning('User "{}" added to the group "{}"'.format(user_id, group_name))
            self.users[user_id] = user_json
            self.users_added.append('{} --> {}'.format(user_id, group_name))

    def log_recap(self):
        """Log info message with an execution recap: a number and a list of changed Dynatrace configurations"""
        recap_str = 'Execution recap\n############################## SUMMARY ########################################'
        nothing_changed = True
        for message, changes_list in [
            ('Created management zones', self.mz_created),
            ('Created user groups', self.groups_created),
            ('Updated user group permissions', self.permissions_updated),
            ('Added users to groups', self.users_added),
        ]:
            changes_num = len(changes_list)
            if changes_num > 0:
                nothing_changed = False
                recap_str += '\n{} ({}):\n'.format(message, changes_num)
                recap_str += '\n'.join(changes_list)
                recap_str += '\n'

        if nothing_changed:
            recap_str += '\nNothing changed in Dynatrace configuration_import_export'

        recap_str += '\n###############################################################################'
        self.log.info(recap_str)

    def assign_roles(self, input_file, users_sheet, roles_sheet, apply_changes=False):
        """Assign Dynatrace roles defined on the roles Excel sheet to the users according to the users sheet table"""
        try:
            self.input_file = input_file
            self.workbook = load_workbook(input_file, read_only=True)
            users_sheet_iterator = self._open_users_sheet(users_sheet)
            self._load_role_definitions(roles_sheet)
            self._get_current_configuration()

            with warnings.catch_warnings():
                # Suppress openpyxl warnings about cell data validation:
                # it is not used in processing, only for convenient file filling
                warnings.simplefilter("ignore")

                for i, row in users_sheet_iterator:
                    # Excel row number is row index + 1
                    row_number = i + 1
                    user = row[0]
                    if not user:
                        # Empty first cell: end of users list
                        break

                    role = row[1]
                    tag_key = row[2]
                    tag_value = row[3]
                    role_level = self._get_role_level(role)
                    group_name = self._get_name_from_row(DtObjectType.GROUP, role, role_level, row_number, row[4:])

                    if role_level == DtRoleLevel.MZ:
                        mz_name = self._get_name_from_row(DtObjectType.MZ, role, role_level, row_number, row[4:])
                        mz_id = self._create_mz_if_needed(mz_name, tag_key, tag_value, apply_changes)
                    else:
                        mz_id = None

                    if group_name in self.groups:
                        group_id = self._update_group_if_needed(group_name, role, role_level, self.env_id, mz_id,
                                                                apply_changes)
                    else:
                        group_id = self._create_group(group_name, role, role_level, self.env_id, mz_id, apply_changes)

                    self._add_user_to_group_if_needed(user, group_id, group_name, apply_changes)

        except IOError as e:
            raise DtRolesError('Could not read the input file "{}": {}'.format(self.input_file, str(e)))


if __name__ == '__main__':
    args = parse_args()
    dt_roles = DynatraceRoles(args.cluster_hostname, args.cluster_token, args.env_id, args.env_token)

    try:
        dt_roles.assign_roles(args.file, args.users_sheet, args.roles_sheet, args.apply_changes)
        dt_roles.log_recap()
        if not args.apply_changes:
            dt_roles.log.warning(
                'Command executed in test mode. ' +
                'To apply changes to Dynatrace configuration_import_export use the flag "--apply-changes".')

    except DtRolesError as e:
        dt_roles.log_recap()
        dt_roles.log.error(str(e))
        exit(1)
